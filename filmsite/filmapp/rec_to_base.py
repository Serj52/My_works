import openpyxl
import os, sys
from sqlite3 import IntegrityError

sys.path.append(r'C:\django_pro\Test\filmsite')
os.environ['DJANGO_SETTINGS_MODULE'] = 'filmsite.settings'
import django


django.setup()
from filmapp.models import Film, Producer
from django.db import IntegrityError

class creat_object_base:

    def __init__(self, name_table):
        self.name = name_table
        wb = openpyxl.load_workbook(r'C:\django_pro\Test\{}'.format(self.name))
        self.sheet = wb.active
        self.rows_max = self.sheet.max_row
        self.cols_max = self.sheet.max_column

    def add_object(self):

        if self.name == 'Режиссеры.xlsx':
            for i in range(2, self.rows_max + 1):
                prod = Producer()
                try:
                    for y in range(1, self.cols_max + 1):
                        title = self.sheet.cell(row=1, column=y).value
                        val = self.sheet.cell(row=i, column=y).value
                        if title == 'producer_name':
                            prod.producer_name = val
                        elif title == 'age':
                            prod.age = val
                        elif title == 'country':
                            prod.country = val
                    prod.save()
                except IntegrityError:
                    continue

        elif self.name == 'Фильмы.xlsx':
            fil = Film()
            for i in range(2, self.rows_max + 1):
                for y in range(1, self.cols_max + 1):
                    title = self.sheet.cell(row=1, column=y).value
                    val = self.sheet.cell(row=i, column=y).value
                    if title == 'name':
                        fil.name = val
                    elif title == 'genre':
                        fil.genre = val
                    elif title == 'year':
                        fil.year = val
                    elif title == 'producer_name':
                        prod = Producer.objects.get(producer_name=val)
                        fil.producer_name_id = prod.id
                fil.save()



if __name__ == '__main__':
    a = creat_object_base('Режиссеры.xlsx')
    a.add_object()


